import pandas as pd
import random
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

ROWS_SOURCE = 5000
ROWS_TARGET = 100

def get_random_date():
    start = datetime(2020, 1, 1)
    end = datetime(2025, 12, 31)
    return start + timedelta(days=random.randint(0, (end - start).days))

def generate_files():
    print("Generiere Testdaten...")

    names = ["Müller", "Schmidt", "Schneider", "Fischer", "Weber", "Meyer", "Wagner", "Becker", "Schulz", "Hoffmann"]
    first_names = ["Max", "Anna", "Tom", "Lisa", "Jan", "Laura", "Tim", "Sarah", "Ben", "Julia"]
    depts = ["IT", "HR", "Sales", "Marketing", "Logistik"]

    data_src = {
        "Kunden_Ref": [f"REF-{i+1000}" for i in range(ROWS_SOURCE)],
        "Kunde_Name": [f"{random.choice(names)}, {random.choice(first_names)}" for i in range(ROWS_SOURCE)],
        "Abteilung_Krz": [random.choice(depts) for _ in range(ROWS_SOURCE)],
        "Transaktionsdatum": [get_random_date() for _ in range(ROWS_SOURCE)],
        "Betrag_Netto": [round(random.uniform(10.0, 5000.0), 2) for _ in range(ROWS_SOURCE)]
    }
    df_src = pd.DataFrame(data_src)
    df_src.to_excel("quelle.xlsx", index=False)
    print(f"-> 'quelle.xlsx' erstellt ({ROWS_SOURCE} Zeilen).")

    data_tgt = {
        "ID": [f"REF-{i}" for i in range(ROWS_TARGET)],
        "Vollständiger Name": [f"{random.choice(names)}, {random.choice(first_names)}" for i in range(ROWS_TARGET)],
        "Datum": [get_random_date() for _ in range(ROWS_TARGET)],
        "Umsatz": [round(random.uniform(100.0, 900.0), 2) for _ in range(ROWS_TARGET)],
        "Steuer (Formel)": [None] * ROWS_TARGET
    }
    df_tgt = pd.DataFrame(data_tgt)
    df_tgt.to_excel("ziel.xlsx", index=False)

    wb = load_workbook("ziel.xlsx")
    ws = wb.active

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)

    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = yellow_fill

    for row in range(2, ROWS_TARGET + 2):
        ws.cell(row=row, column=5).value = f"=D{row}*0.19"

    wb.save("ziel.xlsx")
    print(f"-> 'ziel.xlsx' erstellt ({ROWS_TARGET} Zeilen mit Formatierung & Formeln).")

if __name__ == "__main__":
    generate_files()