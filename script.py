import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from copy import copy
import os
from datetime import datetime

def to_excel_serial(val):
    """Wandelt ein Datum in die Excel-Serial-Nummer (Float) um."""
    if pd.isna(val):
        return None
    # Basis-Datum für Excel (Windows Standard)
    delta = val - datetime(1899, 12, 30)
    return float(delta.days) + (float(delta.seconds) / 86400)

def merge_and_sort_excel(source_path: str, target_path: str):
    if not os.path.exists(source_path) or not os.path.exists(target_path):
        print("Fehler: Dateien nicht gefunden.")
        return

    # --- 1. Analyse der Zieldatei ---
    print("Lade Zieldatei...")
    wb = load_workbook(target_path)
    ws = wb.active

    target_headers = {cell.value: cell.column for cell in ws[1] if cell.value}
    header_names = list(target_headers.keys())
    
    # Template-Infos sichern (Formatierung & Formeln)
    template_info = {} 
    if ws.max_row >= 2:
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=2, column=col_idx)
            template_info[col_idx] = {
                'font': copy(cell.font),
                'border': copy(cell.border),
                'fill': copy(cell.fill),
                'number_format': copy(cell.number_format), # WICHTIG: Hier steht z.B. "DD.MM.YYYY"
                'alignment': copy(cell.alignment),
                'protection': copy(cell.protection),
                'is_formula': cell.data_type == 'f',
                'formula': cell.value if cell.data_type == 'f' else None,
                'col_letter': cell.column_letter
            }

    # Bestandsdaten laden
    df_target_old = pd.read_excel(target_path)
    df_target_old = df_target_old.loc[:, ~df_target_old.columns.str.contains('^Unnamed')]

    # --- 2. Quelldaten laden & Mapping ---
    print(f"Lade Quelldatei '{source_path}'...")
    df_source = pd.read_excel(source_path)

    print(f"\n--- Spaltenzuordnung ---")
    mapping = {}
    for t_col in header_names:
        col_idx = target_headers[t_col]
        # Formel-Spalten überspringen
        if template_info.get(col_idx, {}).get('is_formula'):
            continue

        s_col = input(f"Quelle für Ziel '{t_col}'? (Enter für leer): ").strip()
        if s_col in df_source.columns:
            mapping[t_col] = s_col

    # DataFrame zusammenbauen
    df_new_data = pd.DataFrame()
    for t_col in header_names:
        if t_col in mapping:
            df_new_data[t_col] = df_source[mapping[t_col]]
        else:
            df_new_data[t_col] = None

    # --- 3. Merge & Clean Up Dates ---
    df_target_old = df_target_old[header_names] if not df_target_old.empty else pd.DataFrame(columns=header_names)
    df_total = pd.concat([df_target_old, df_new_data], ignore_index=True)

    # DATUM KONVERTIERUNG (Fix für Type Error)
    print("Konvertiere Datumsformate in Excel-Zahlen...")
    for col in df_total.columns:
        # Prüfen, ob Spalte wie ein Datum aussieht (datetime64)
        if pd.api.types.is_datetime64_any_dtype(df_total[col]):
            # Umwandlung: Timestamp -> Python Datetime -> Excel Serial Float
            df_total[col] = df_total[col].apply(lambda x: to_excel_serial(pd.to_datetime(x)) if pd.notnull(x) else None)

    # --- 4. Sortieren ---
    print("\n--- Sortierung ---")
    sort_col = input(f"Nach welcher Ziel-Spalte sortieren? ({', '.join(header_names)}): ").strip()
    
    if sort_col in df_total.columns:
        print(f"Sortiere {len(df_total)} Zeilen nach '{sort_col}'...")
        try:
            # Versuch 1: Normale Sortierung (korrekt für reine Zahlen/Datumsangaben)
            df_total = df_total.sort_values(by=sort_col)
        except TypeError:
            # Fallback: Bei gemischten Typen (int vs str) als String sortieren, um Crash zu verhindern
            print(f"Warnung: Spalte '{sort_col}' enthält gemischte Datentypen. Sortiere als Text.")
            df_total = df_total.sort_values(by=sort_col, key=lambda x: x.astype(str))
    
    # --- 5. Schreiben ---
    print("Schreibe Daten...")
    start_row = 2
    
    for i, (_, row_data) in enumerate(df_total.iterrows()):
        current_row = start_row + i
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            header_name = ws.cell(row=1, column=col_idx).value
            t_info = template_info.get(col_idx)
            
            if not t_info: continue

            # A) Inhalt
            if t_info['is_formula'] and t_info['formula']:
                origin_coord = f"{t_info['col_letter']}2"
                target_coord = f"{t_info['col_letter']}{current_row}"
                cell.value = Translator(t_info['formula'], origin=origin_coord).translate_formula(target_coord)
            else:
                val = row_data.get(header_name)
                # NaN/None Behandlung
                if pd.isna(val) or val is None:
                    cell.value = None
                else:
                    cell.value = val 
                    # Hier wird jetzt ein Float geschrieben, falls es ein Datum war

            # B) Styles (Hier wird das Datumsformat "DD.MM.YYYY" wiederhergestellt)
            if t_info['number_format']:
                cell.number_format = copy(t_info['number_format'])
            cell.font = copy(t_info['font'])
            cell.border = copy(t_info['border'])
            cell.fill = copy(t_info['fill'])
            cell.alignment = copy(t_info['alignment'])
            cell.protection = copy(t_info['protection'])

    try:
        wb.save(target_path)
        print(f"Erfolgreich gespeichert! ({len(df_total)} Zeilen)")
    except PermissionError:
        print("Fehler: Datei ist geöffnet.")

if __name__ == "__main__":
    s = input("Quell-Datei: ").strip().strip('"')
    t = input("Ziel-Datei: ").strip().strip('"')
    merge_and_sort_excel(s, t)